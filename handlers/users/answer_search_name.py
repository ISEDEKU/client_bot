import os
import re
import urllib

import openpyxl
from aiogram import types
from aiogram.dispatcher import FSMContext
from openpyxl.utils import get_column_letter

from keyboards.default.main_keyboard import menu_1
from loader import dp
from states.state_class import Questions


@dp.message_handler(state=Questions.search_name)  # хэндлер, с первым состоянием, который принимает ответ
async def answer_search_name(message: types.Message, state=FSMContext):
    await message.answer('Ищу все товары с похожим названием...')
    urllib.request.urlretrieve('http://aparat.ua/cli/telegram_opt.xlsx', 'out.xlsx')
    path_to_file = os.path.abspath('out.xlsx')
    book = openpyxl.open(path_to_file, read_only=True)  # открываем книгу, в аргументе указываем путь к файлу
    sheet = book.active  # получаем первый и единственный лист

    sheet_cells = []

    answer = str(message.text)  # в переменную answer помещяется то, что ответил пользователь в ТГ

    search_text = str(answer)  # переменная, хранит в себе значение, которое мы ищём

    wb = openpyxl.load_workbook(path_to_file)  # Грузим наш прайс-лист
    sheets_list = wb.sheetnames  # Получаем список всех листов в файле
    sheet_active = wb[sheets_list[0]]  # Начинаем работать с самым первым
    row_max = sheet_active.max_row  # Получаем количество столбцов
    # print(type(row_max))
    column_max = sheet_active.max_column  # Получаем количество строк

    row_min = 1  # Переменная, отвечающая за номер строки
    column_min = 1  # Переменная, отвечающая за номер столбца

    while column_min <= column_max:
        row_min_min = row_min
        row_max_max = row_max
        while row_min_min <= row_max_max:
            row_min_min = str(row_min_min)

            word_column = get_column_letter(column_min)
            word_column = str(word_column)
            word_cell = word_column + row_min_min
            data_from_cell = sheet_active[word_cell].value
            data_from_cell = str(data_from_cell)
            regular = search_text
            result = re.findall(regular.lower(), data_from_cell.lower())

            if len(result) > 0:
                sheet_cells.append(word_cell)
            # каждая ячейка, содержащая в себе нужное нам значение, помещается в лист

            row_min_min = int(row_min_min)
            row_min_min = row_min_min + 1
        column_min = column_min + 1
        # алгоритм для сравнения ответа из ТГ и товарами в exel файле

    await state.update_data(answer1=sheet_cells)
    # помещаем в хранилище лист, содержащий в себе адреса всех нужных нам ячеек

    if len(sheet_cells) == 0:
        await message.answer('Оу, я не нашёл ничего подобного...\nНапишите другое название:')
        await state.reset_state()
        await Questions.search_name.set()

    name_list = []

    if len(sheet_cells) > 1:
        i = 1  # добавлю счётчик
        for name in sheet_cells:
            if len(name) == 2:
                name = int(name[1])
                name_list.append((f'{sheet[name][1].value} --- {[i]}'))
                # await message.answer(f'{sheet[name][1].value} --- {[i]}')
            elif len(name) > 2:
                name = name[1:]
                name = int(name)
                name_list.append((f'{sheet[name][1].value} --- {[i]}'))
                # await message.answer(f'{sheet[name][1].value} --- {[i]}')
            i += 1
            await state.update_data(answer2=i)
            if i > 100:
                await message.answer('Оу... Список слишком большой, введите название поточнее:')
                await state.reset_state()
                await Questions.search_name.set()
                return

        await message.answer(("\n\n".join(name_list)))
        # алгоритм который берёт ячейки из списка, и выдаёт всю нужную информацию в строке с этой ячейкой
        await message.answer(f'Я нашел {len(sheet_cells)} товаров с похожим названием!\nВведите нужный номер:')
        await Questions.serial_num.set()  # запуск второго состояния

    elif len(sheet_cells) == 1:
        sel_num = sheet_cells[0]
        if len(sel_num) == 2:
            sel_num = int(sel_num[1])
            product_name = sheet[sel_num][1].value
            quantity = sheet[sel_num][3].value
            cost_grn = sheet[sel_num][7].value
            if int(quantity) < 3 or quantity == ' ':
                quantity = 'Уточните у менеджера'
            else:
                quantity = 'Есть в наличии'
            # # алгоритм который берёт ячейку из списка, и выдаёт всю нужную информацию в строке с этой ячейкой
            await message.answer(
                f'\n{product_name}\nНаличие товара:  {quantity}\nЦена на сайте:  {cost_grn} грн.'
            )

            await message.answer('Поиск завершён!', reply_markup=menu_1)
            await state.finish()

        elif len(sel_num) > 2:
            sel_num = sel_num[1:]
            sel_num = int(sel_num)
            product_name = sheet[sel_num][1].value
            quantity = sheet[sel_num][3].value
            cost_grn = sheet[sel_num][7].value
            if int(quantity) < 3 or quantity == ' ':
                quantity = 'Уточните у менеджера'
            else:
                quantity = 'Есть в наличии'
            # алгоритм который берёт ячейку из списка, и выдаёт всю нужную информацию в строке с этой ячейкой
            await message.answer(
                f'\n{product_name}\nНаличие товара:  {quantity}\nЦена на сайте:  {cost_grn} грн.'
            )
